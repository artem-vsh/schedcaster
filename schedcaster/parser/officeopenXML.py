# -*- coding: utf-8 -*-
"""
Created on Thu Aug  2 19:16:58 2012

@author: avsh
"""

from openpyxl import load_workbook
import schedcaster
import hashlib


def parse(filename, spec):
    wb = load_workbook(filename)
    return parseWorkbook(wb, spec)


def parseWorkbook(wb, spec):
    entries = []

    # iterate only sheets, that we support via sheetToType
    for sheetName in filter(lambda s: s in spec.sheets, wb.get_sheet_names()):
        sheet = wb.get_sheet_by_name(sheetName)
        sheetSpec = spec.sheets[sheetName]

        rows = sheet.get_highest_row()
        for i in range(1, rows + 1):
            rowName = str(i)
            entry = schedcaster.scheduler.Entry(
                                    state=schedcaster.scheduler.STATE_ONESHOT,
                                    handler='post')

            poisoned = False  # indicates if row cannot be parsed
            for col in sheetSpec.columns.items():
                columnName = col[0]
                argName = col[1].argName
                filterFn = col[1].filterFn
                # use trivial preprocessor if none given
                preprocessorFn = col[1].processorFn

                cellValue = sheet.cell(columnName + rowName).value
                # if filter fails, skip this row
                if not filterFn(cellValue):
                    poisoned = True
                    break
                cellRealValue = preprocessorFn(cellValue)

                if argName[0] == '@':
                    argName = argName[1:]
                    if argName == 'cron':
                        entry.cron = cellRealValue
                    else:
                        raise RuntimeError('wrong parameter @%s' % columnName)
                else:
                    entry.arg(argName, cellRealValue)

            # if row has incorrect format, do not add it
            if poisoned:
                continue

            # compute hash based on the cells provided
            # hash is used to check if the value is already used in the db
            hashSrc = "&".join(
                map(lambda col: str(sheet.cell(col + rowName).value or ""),
                    sheetSpec.hashSpec.columnNames)).encode('utf-16')
            entry.arg('hash', hashlib.md5(hashSrc).digest())
            #use as id instead
            entry.id = hashSrc

            entries.append(entry)

    return entries


class Spec(object):
    def __init__(self, sheets=[]):
        self.sheets = {}
        for sheet in sheets:
            self.sheets[sheet.name] = sheet


class SheetSpec(object):
    def __init__(self, name, columns=[], hashSpec=None):
        self.name = name
        self.columns = {}
        for column in columns:
            self.columns[column.name] = column
        self.hashSpec = hashSpec


class ColumnSpec(object):
    def __init__(self, name, argName,
                      filterFn=lambda v: True,
                      processorFn=lambda v: v):
        self.name = name
        self.argName = argName
        self.filterFn = filterFn
        self.processorFn = processorFn


class HashSpec(object):
    def __init__(self, columnNames=[]):
        self.columnNames = columnNames
